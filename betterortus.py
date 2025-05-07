# A Project for finding marks at ORTUS automatization
# When it works:
#   Script only works if ORTUS is in Latvian
#   You have to give script valid username and password
#
# What it does:
#   Script gets current/last semester marks from page
#   Saves data in excel format

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

from typing import List, Tuple
from collections import defaultdict
import getpass

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

CourseMap = defaultdict[str, List[Tuple[str, str]]]
courseMap: CourseMap = defaultdict(list) # Global container which holds course name and its grades
semesterName  = ""                       # What year courses we will collect // Exported to excel
loginUsername = ""                       # To login in ortus
loginPassword = ""                       # To login in ortus

# Function which searches ORTUS for marks
# Returns False if search was unsuccessful
def HandleWebSearch():
    global semesterName

    chrome_options = Options()
    chrome_options.add_argument("--headless")               # Run Chrome in headless mode
    chrome_options.add_argument("--disable-gpu")            # Recommended for Windows

    driver = webdriver.Chrome(options=chrome_options)

    # Login in ortus
    driver.get("https://id2.rtu.lv/openam/UI/Login?module=LDAP&locale=lv")
    driver.find_element(By.ID, "IDToken1").send_keys(loginUsername)
    driver.find_element(By.ID, "IDToken2").send_keys(loginPassword)

    old_url = driver.current_url

    driver.find_element(By.NAME, "Login.Submit").click()

    # Wait for redirect // ortus does that for some reason // ortus is pretty strange when you check its code
    WebDriverWait(driver, 10).until(EC.url_changes(old_url))

    # Check if login was successful
    if "openam" in driver.current_url.lower():
        print("Failed to login in ORTUS, check credentials")
        return False

    # Find link to semester page and click on it
    driver.find_element(By.ID, "portalNavigationTabGroupsList").find_element(By.XPATH, './/a[@title="Studentiem"]').click()

    # This is where latest semester courses are at
    semesterCourses = driver.find_element(By.CLASS_NAME, "moodle-courses").find_elements(By.XPATH, './*')

    semesterName   = semesterCourses[1].text.strip()                        # global name 
    linksToCourses = semesterCourses[2].find_elements(By.TAG_NAME, "a")     # Links to all courses in semester

    print(f"\n\nYear: {semesterName}\n")

    # Iterate thru all courses
    for link in linksToCourses:
        courseName = link.text.strip()  # What course is it? // Exported to excel
        href       = link.get_attribute("href")
        print(f" - {courseName}: {href}") # where driver is going to

        driver.get(href)

        try:
            # Navigate to marks page
            marksAside  = driver.find_element(By.ID, "block-region-side-post") # This is where link to marks live in
            linkToMarks = marksAside.find_elements(By.TAG_NAME, "a")
            driver.get(linkToMarks[1].get_attribute("href"))                   # Aside have multiple links, second one has link to marks

            try:
                tableRows = driver.find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, "tr") 

                for row in tableRows:
                    # First item contains title
                    # Third row contains value
                    rowItems = row.find_elements(By.TAG_NAME, "td")

                    markTitle = "" # Exported to excel
                    markValue = "" # Exported to excel

                    if len(rowItems) >= 3: # If there are less, then row is not mark row
                        try:
                            markTitle = rowItems[0].find_element(By.TAG_NAME, "a").text.strip()
                            markValue = rowItems[2].text.strip()
                        except:
                            continue

                    if markTitle != "":
                        courseMap[courseName].append((markTitle, markValue))

            except: # Happens because sometimes rows are randomly completely empty
                print("### Could not read from table, invalid row design")
                continue

        except: # If link somehow was invalid
            print(f"### Could not open link to {linkToMarks[1]}")
            continue

        print(f"Got results in {courseName}")

        driver.back()  # from marks page to course page
        driver.back()  # from course page to course list

    return True


# Exports all collected data in "courseMap" after web search to excel
def ExportToExcel():
    global semesterName

    wb = Workbook()
    ws = wb.active

    ws.title = semesterName
    
    # Setup Headers and style
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 20
    headers = ["Kurss", "Pārbaudījums", "Atzīme"]
    ws.append(headers)
    for colNum, header in enumerate(headers, 1):
        cell      = ws.cell(row=1, column=colNum)
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type='solid')
        cell.font = Font(color="FFFFFF", bold=True)

    row = 2 # Start writing after header
    for course, exercises in courseMap.items():
        for exercise, grade in exercises:
            ws.cell(row=row, column=1).value = course
            ws.cell(row=row, column=2).value = exercise
            ws.cell(row=row, column=3).value = grade
            row += 1
        row += 1 # Between every course give one row gap

    wb.save("atzimes.xlsx")

    print("Excel is now created!")

def PrintBanner():
    try:
        with open("banner.txt", "r", encoding="utf-8") as f:
            print(f.read())
    except:
        print("Why would you delete banner?")

if __name__ == "__main__":
    print("\033[H\033[2J", end="") # To start program on clean screen
    PrintBanner()

    loginUsername = input("Enter ORTUS username: ")
    loginPassword = getpass.getpass("Enter ORTUS password: ")

    if not HandleWebSearch(): # Continue only if getting data was successful
        exit()

    ExportToExcel()